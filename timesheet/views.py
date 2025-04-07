from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.timezone import now
from django.http import HttpResponse
import pandas as pd
from django.db.models import Count
from django.contrib.auth.models import User
from django.contrib import messages
import openpyxl
from io import BytesIO
from datetime import datetime
from .models import Employee, EmployeeTimesheet
from django.http import JsonResponse
from django.utils import timezone
from django.db import models


def index(request):

    return render(request,'index.html')

# Check if user is admin
def is_admin(user):
    return user.is_superuser

# Manual User Registration View
def register(request):
    if request.method == "POST":
        username = request.POST.get('username')
        employee_id = request.POST.get('employee_id')
        password = request.POST.get('password')
        department = request.POST.get('department')
        role = request.POST.get('role')
        fullname = request.POST.get('fullname')

        if Employee.objects.filter(employee_id=employee_id).exists():
            messages.error(request, "Employee ID already exists!")
            return redirect('register')

        
        user = Employee.objects.create_user(username=username,
            fullname = fullname,
            employee_id=employee_id,
            department=department,
            role=role,
            password=password)
        
        user.save()
    
    return render(request, 'register.html')
from django.contrib.auth.hashers import check_password  # Import check_password


from django.utils import timezone  # Import Django's timezone
from datetime import timedelta
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import check_password
from .models import Employee, EmployeeTimesheet

def sign_in(request):
    if request.method == "POST":
        employee_id = request.POST['employee_id']
        password = request.POST['password']
        check_in_time = timezone.now()  # Use timezone-aware datetime

        try:
            # Retrieve the employee by their employee ID
            employee = Employee.objects.get(employee_id=employee_id)

            # Check if the entered password matches the stored password
            if check_password(password, employee.password):
                today = check_in_time.date()

                # Get the most recent sign-in entry for this employee
                last_entry = EmployeeTimesheet.objects.filter(
                    employee=employee
                ).order_by('-check_in_time').first()

                if last_entry:
                    time_difference = check_in_time - last_entry.check_in_time
                    
                    # Check if they signed in but never signed out
                    if last_entry.check_out_time is None:
                        if time_difference.total_seconds() < 43200:  # Less than 12 hours
                            messages.error(request, "You must sign out first before signing in again.")
                            return redirect('sign_in')
                    else:
                        # If they signed out, allow new sign-in
                        pass

                # **Delete yesterday's entry if they didn't sign out**
                yesterday = today - timedelta(days=1)
                EmployeeTimesheet.objects.filter(
                    employee=employee,
                    check_in_time__date=yesterday,
                    check_out_time__isnull=True  # Only remove entries with no sign-out
                ).delete()

                # **Save new sign-in**
                EmployeeTimesheet.objects.create(employee=employee, check_in_time=check_in_time)
                messages.success(request, "Sign-in successful!")
                return redirect('sign_in')  # Redirect to the main dashboard

            else:
                messages.error(request, "Incorrect password. Please try again.")
                return redirect('sign_in')

        except Employee.DoesNotExist:
            messages.error(request, "Employee ID not found. Please try again.")
            return redirect('sign_in')

    return render(request, 'sign_in.html')

def sign_out(request):
    if request.method == "POST":
        employee_id = request.POST['employee_id']
        password = request.POST['password']
        check_out_time = now()  # Use the current time for check-out time

        try:
            # Retrieve the employee by their employee ID
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Check if the entered password matches the stored password
            if check_password(password, employee.password):
                # Retrieve the latest timesheet entry where the employee checked in
                timesheet = EmployeeTimesheet.objects.filter(employee=employee, check_in_time__isnull=False).last()

                if timesheet:
                    last_check_in = timesheet.check_in_time
                    
                    # Check if the last check-in was more than 15 hours ago
                    if last_check_in and (check_out_time - last_check_in > timedelta(hours=15)):
                        error_message = "Your last sign-in was more than 15 hours ago. Please sign in again before signing out."
                        return render(request, 'sign_out.html', {'error_message': error_message})
                    
                    # Update the existing timesheet with the check_out_time
                    timesheet.check_out_time = check_out_time
                    timesheet.save()
                    return redirect('index')  # Redirect to a dashboard or another page on successful sign-out
                else:
                    # If no matching check-in time is found
                    error_message = "No check-in record found. Please check if you have signed in first."
                    return render(request, 'sign_out.html', {'error_message': error_message})
            else:
                # If password is incorrect, show an error message
                error_message = "Incorrect password. Please try again."
                return render(request, 'sign_out.html', {'error_message': error_message})
        except Employee.DoesNotExist:
            # If the employee ID doesn't exist in the database
            error_message = "Employee ID not found. Please try again."
            return render(request, 'sign_out.html', {'error_message': error_message})

    return render(request, 'sign_out.html')

def login_view(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # If user is admin, redirect to admin dashboard
            if user.is_superuser:
                return redirect('admin_dashboard')  # Redirect to admin dashboard
            else:
                return redirect('index')  # Regular user redirection
        
        else:
            error_message = "Invalid username or password."
            return render(request, 'login.html', {'error_message': error_message})
    
    return render(request, 'login.html')

# Admin Dashboard (Summary)
@user_passes_test(is_admin)
def admin_dashboard(request):
    total_employees = Employee.objects.count()
    employees_on_site = EmployeeTimesheet.objects.filter(check_out_time__isnull=True).count()
    employees_per_department = Employee.objects.values('department').annotate(count=Count('department'))
    
    context = {
        'total_employees': total_employees,
        'employees_on_site': employees_on_site,
        'employees_per_department': employees_per_department
    }
    return render(request, 'admin_dashboard.html', context)



def get_employee_attendance_data(request):
    # Get the current date (today's date)
    today = timezone.now().date()  # We use .date() to ignore the time part
    
    # Get the number of check-ins for each department today
    check_ins_by_dept = EmployeeTimesheet.objects.filter(
        check_in_time__date=today  # Filter only today's date
    ).values('employee__department').annotate(check_in_count=models.Count('employee'))

    # Get the total number of check-ins for today
    total_check_ins = EmployeeTimesheet.objects.filter(
        check_in_time__date=today
    ).count()

    # Get the total number of check-ins for today
    total_check_out = EmployeeTimesheet.objects.filter(
        check_out_time__date=today
    ).count()
    active_employees = total_check_ins - total_check_out

    # Create a dictionary of department and check-in counts
    department_data = {entry['employee__department']: entry['check_in_count'] for entry in check_ins_by_dept}

    # Return the data as JSON, including the total check-ins for today
    return JsonResponse( {
        'attendance_data': department_data,
        'total_check_ins': total_check_ins,
        'total_check_out': total_check_out,
        'active_employees': active_employees
    })

from datetime import timedelta
@user_passes_test(is_admin)
def view_sheet(request):
    # Get today's date
    today = timezone.now().date()

    # Get the selected time range from the query parameter (default to "today")
    time_range = request.GET.get('time_range', 'today')

    # Get the date range based on the selected filter
    if time_range == 'today':
        start_date = today
        end_date = today
    elif time_range == 'one_week':
        start_date = today - timedelta(weeks=1)
        end_date = today
    elif time_range == 'one_month':
        start_date = today - timedelta(weeks=4)
        end_date = today
    elif time_range == 'one_year':
        start_date = today - timedelta(days=365)
        end_date = today
    elif time_range == 'all':
        # No filtering, fetch all timesheets
        timesheets = EmployeeTimesheet.objects.all().select_related('employee')
        return render(request, 'view_sheet.html', {'timesheets': timesheets, 'time_range': time_range})
    else:
        # Default to today if no valid filter is provided
        start_date = today
        end_date = today

    # Fetch employee timesheets for the selected date range
    timesheets = EmployeeTimesheet.objects.filter(check_in_time__date__range=[start_date, end_date]).select_related('employee')

    return render(request, 'view_sheet.html', {'timesheets': timesheets, 'time_range': time_range})

import datetime
import openpyxl
from django.http import HttpResponse

def get_month_week(date):
    """Determine the correct month and week, ensuring no 5th week exists and each month starts fresh."""
    if date.year == 2025 and date.month == 1:
        first_week_start = datetime.date(2025, 1, 7)
    else:
        first_week_start = date.replace(day=1)  

    if date.day < 7:
        return date.strftime("%B"), 1

    week_number = ((date - first_week_start).days // 7) + 1

    if week_number > 4:
        next_month = (date.month % 12) + 1
        next_year = date.year if date.month < 12 else date.year + 1
        return datetime.date(next_year, next_month, 1).strftime("%B"), 1

    return date.strftime("%B"), week_number



def download_timesheets(request):
    search_query = request.GET.get("search", "").strip()
    time_range = request.GET.get("time_range", "")

    today = datetime.date.today()
    start_date, end_date = today, today

    if time_range == "one_month":
        start_date = today - datetime.timedelta(days=30)
    elif time_range == "one_week":
        start_date = today - datetime.timedelta(days=7)
    elif time_range == "one_year":
        start_date = today - datetime.timedelta(days=360)

    month_name, week_number = get_month_week(start_date)
    sheet_name = f"{month_name} Week {week_number}"

    file_path = "timesheet.xlsx"
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)  # Clear existing data
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Insert Date Range at the Top
    ws["A1"] = "Date From:"
    ws["B1"] = start_date.strftime("%Y-%m-%d")
    ws["A2"] = "Date To:"
    ws["B2"] = end_date.strftime("%Y-%m-%d")

    headers = [
        "NO.", "Employee ID", "Full Name", "Working Days",
        "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday",
        "Total Hours"
    ]

    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_num, value=header)

    timesheets = EmployeeTimesheet.objects.filter(
        employee__employee_id__icontains=search_query,
        check_in_time__date__gte=start_date,
        check_in_time__date__lte=end_date
    ).order_by("check_in_time")

    employee_data = {}
    for timesheet in timesheets:
        emp_id = timesheet.employee.employee_id
        full_name = timesheet.employee.fullname
        check_in_date = timesheet.check_in_time.date()
        hours = timesheet.hours_worked or 0

        if emp_id not in employee_data:
            employee_data[emp_id] = {
                "full_name": full_name,
                "shift_days": 0,
                "daily_hours": {day: 0 for day in ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]},
                "total_hours": 0
            }

        day_name = check_in_date.strftime("%A")
        employee_data[emp_id]["daily_hours"][day_name] += hours
        employee_data[emp_id]["shift_days"] += 1
        employee_data[emp_id]["total_hours"] += hours

    row_num = 4  
    total_shift_days = 0
    total_hours_per_day = {day: 0 for day in ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]}
    total_hours_overall = 0

    for index, (emp_id, data) in enumerate(employee_data.items(), start=1):
        row_data = [
            index,
            emp_id,
            data["full_name"],
            data["shift_days"],
            data["daily_hours"]["Sunday"],
            data["daily_hours"]["Monday"],
            data["daily_hours"]["Tuesday"],
            data["daily_hours"]["Wednesday"],
            data["daily_hours"]["Thursday"],
            data["daily_hours"]["Friday"],
            data["daily_hours"]["Saturday"],
            round(data["total_hours"], 2)
        ]

        total_shift_days += data["shift_days"]
        total_hours_overall += data["total_hours"]

        for day in total_hours_per_day:
            total_hours_per_day[day] += data["daily_hours"][day]

        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

        row_num += 1  

    # Insert total row
    total_row = [
        "TOTAL", "", "", total_shift_days,
        total_hours_per_day["Sunday"],
        total_hours_per_day["Monday"],
        total_hours_per_day["Tuesday"],
        total_hours_per_day["Wednesday"],
        total_hours_per_day["Thursday"],
        total_hours_per_day["Friday"],
        total_hours_per_day["Saturday"],
        round(total_hours_overall, 2)
    ]

    # Insert total hours overall
    ws.cell(row=row_num, column=11, value="Total Hours Overall:").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=row_num, column=12, value=round(total_hours_overall, 2)).font = openpyxl.styles.Font(bold=True)


    wb.save(file_path)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{file_path}"'
    wb.save(response)
    return response
    
from django.contrib.auth import logout

def logout_view(request):
    logout(request)  # This will log out the user
    response = redirect('login')  # Redirect to the login page

    # Prevent the browser from caching the page to avoid back-button navigation
    response['Cache-Control'] = 'no-store'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'

    return response
def get_email(request):

    return render(request,'get_email.html')

def check_email_exists(request):
    if request.method == "POST":
        email = request.POST.get("email")
        if Employee.objects.filter(username=email).exists():
            return JsonResponse({"exists": True})
        else:
            return JsonResponse({"exists": False})
    return JsonResponse({"error": "Invalid request"}, status=400)

from django.contrib.auth.hashers import make_password
def reset_password(request):
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        # Check if employee ID exists in the database
        try:
            user = Employee.objects.get(employee_id=employee_id)  # Fetch user by employee_id
        except Employee.DoesNotExist:
            messages.error(request, "Employee ID not found.")
            return redirect("reset-password")

        # Validate passwords
        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect("reset-password")

        # Update the user's password
        user.password = make_password(new_password)  # Hash the password
        user.save()

        messages.success(request, "Password reset successfully!")
        return redirect("reset-password")

    return render(request, "reset-password.html")


#face id implementation
from deepface import DeepFace
import os
from django.conf import settings

def register_face(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        face_photo = request.FILES.get('face_photo')
        
        try:
            employee = Employee.objects.get(employee_id=employee_id)
            
            # Save the face photo
            employee.face_photo = face_photo
            employee.save()
            
            messages.success(request, "Face registered successfully!")
            return redirect('index')
            
        except Employee.DoesNotExist:
            messages.error(request, "Employee not found")
    
    return render(request, 'register_face.html')
import time
def face_sign_in(request):
    if request.method == 'POST':
        if 'face_photo' not in request.FILES:
            messages.error(request, "Please capture or upload a face photo")
            return redirect('face_sign_in')

        uploaded_file = request.FILES['face_photo']

        if uploaded_file.size > 5 * 1024 * 1024:
            messages.error(request, "Image too large (max 5MB)")
            return redirect('face_sign_in')

        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp')
        os.makedirs(temp_dir, exist_ok=True)

        temp_filename = f"temp_face_{request.session.session_key}_{int(time.time())}.jpg"
        temp_path = os.path.join(temp_dir, temp_filename)

        try:
            with open(temp_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            employees = Employee.objects.filter(face_photo__isnull=False)

            verification_params = {
                'model_name': 'Facenet',
                'detector_backend': 'opencv',
                'distance_metric': 'cosine',
                'enforce_detection': True,
                'threshold': 0.65
            }

            match_found = False

            for employee in employees:
                try:
                    print(f"Checking face for: {employee.employee_id}")
                    result = DeepFace.verify(
                        img1_path=temp_path,
                        img2_path=employee.face_photo.path,
                        **verification_params
                    )
                    print(f"Verification result for {employee.employee_id}: {result}")

                    if result['verified']:
                        match_found = True
                        login(request, employee)
                        check_in_time = timezone.now()
                        today = check_in_time.date()

                        last_entry = EmployeeTimesheet.objects.filter(
                            employee=employee
                        ).order_by('-check_in_time').first()

                        if last_entry:
                            time_difference = check_in_time - last_entry.check_in_time
                            if last_entry.check_out_time is None:
                                if time_difference.total_seconds() < 43200:
                                    messages.error(request, "You must sign out first before signing in again.")
                                    return redirect('face_sign_in')

                        yesterday = today - timedelta(days=1)
                        EmployeeTimesheet.objects.filter(
                            employee=employee,
                            check_in_time__date=yesterday,
                            check_out_time__isnull=True
                        ).delete()

                        EmployeeTimesheet.objects.create(
                            employee=employee,
                            check_in_time=check_in_time
                        )

                        messages.success(request, f"Welcome back, {employee.fullname}!")
                        return redirect('index')

                except Exception as e:
                    print(f"Face comparison error for {employee.employee_id}: {str(e)}")
                    continue

            if not match_found:
                print("No verified face match found.")
                messages.error(request, "Face not recognized. Please try again or use password login.")

        except Exception as e:
            messages.error(request, "Error processing your photo")
            print(f"Face auth system error: {str(e)}")

        finally:
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except:
                pass

    return render(request, 'signin_face.html')


def face_sign_out(request):
    if request.method == 'POST':
        if 'face_photo' not in request.FILES:
            messages.error(request, "Please capture or upload a face photo")
            return redirect('face_sign_out')

        uploaded_file = request.FILES['face_photo']

        if uploaded_file.size > 5 * 1024 * 1024:
            messages.error(request, "Image too large (max 5MB)")
            return redirect('face_sign_out')

        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp')
        os.makedirs(temp_dir, exist_ok=True)

        temp_filename = f"temp_face_{request.session.session_key}_{int(time.time())}.jpg"
        temp_path = os.path.join(temp_dir, temp_filename)

        try:
            with open(temp_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            employees = Employee.objects.filter(face_photo__isnull=False)

            verification_params = {
                'model_name': 'Facenet',
                'detector_backend': 'opencv',
                'distance_metric': 'cosine',
                'enforce_detection': True,
                'threshold': 0.65
            }

            match_found = False

            for employee in employees:
                try:
                    print(f"Checking face for: {employee.employee_id}")
                    result = DeepFace.verify(
                        img1_path=temp_path,
                        img2_path=employee.face_photo.path,
                        **verification_params
                    )
                    print(f"Verification result for {employee.employee_id}: {result}")

                    if result['verified']:
                        match_found = True
                        logout(request)
                        check_out_time = timezone.now()
                        today = check_out_time.date()

                        # Get the last sign-in entry for this employee
                        last_entry = EmployeeTimesheet.objects.filter(
                            employee=employee, check_in_time__isnull=False
                        ).order_by('-check_in_time').first()

                        if last_entry:
                            last_check_in = last_entry.check_in_time
                            time_difference = check_out_time - last_check_in

                            # Check if the last check-in was more than 15 hours ago
                            if time_difference > timedelta(hours=15):
                                messages.error(request, "Your last sign-in was more than 15 hours ago. Please sign in again before signing out.")
                                return redirect('face_sign_out')

                            # Record the check-out time if sign-out is allowed
                            last_entry.check_out_time = check_out_time
                            last_entry.save()

                            messages.success(request, f"Goodbye, {employee.fullname}! You've successfully signed out.")
                            return redirect('index')

                        else:
                            messages.error(request, "No valid check-in record found for sign-out.")
                            return redirect('face_sign_out')

                except Exception as e:
                    print(f"Face comparison error for {employee.employee_id}: {str(e)}")
                    continue

            if not match_found:
                messages.error(request, "Face not recognized. Please try again or use password login.")

        except Exception as e:
            messages.error(request, "Error processing your photo")
            print(f"Face auth system error: {str(e)}")

        finally:
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except:
                pass

    return render(request, 'signout_face.html')
